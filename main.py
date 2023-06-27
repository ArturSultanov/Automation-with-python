# Artur Sultanov June 2023
import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]


products_num_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

# Task 1 - calculation number of product per supplier
    if supplier_name in products_num_per_supplier:
        products_num_per_supplier[supplier_name] += 1
    else:
        products_num_per_supplier[supplier_name] = 1

# Task 2 - calculation total value of inventory per supplier (inventory * price)
    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] += inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

# Task 3 - logic products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

# Task 4 - add value for total inventory price
    inventory_price.value = inventory * price

inv_file.save("inventory_with_total_value.xlsx")

# print(products_num_per_supplier)
# print(total_value_per_supplier)
# print(products_under_10_inv)
